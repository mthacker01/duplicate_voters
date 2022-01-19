import pandas as pd
from rich import print
from openpyxl import load_workbook
from openpyxl.styles import Font

master_path = input(str('Enter file path of master file:  '))

if master_path.endswith('.xlsx'):
    pass
else:
    print('File path must end with ".xlsx".')
    input(str('Enter file path of master file:  '))

to_path = input('Path to store county files:  ')

master_df = pd.read_excel(master_path)


for county in range(1, 121):
    primary_county = master_df[(master_df['COUNTY'] == county)]
    duplicate_county = master_df[(master_df['COUNTY 2'] == county) & (master_df['COUNTY'] != county) ]
    final = primary_county.append(duplicate_county)

    final.to_excel(f'{to_path}/{county}.xlsx', index=False)

    wb = load_workbook(f'{to_path}/{county}.xlsx')

    sheet = wb.active

    total_rows = sheet.max_row

    for row in sheet[f'T1:AK{total_rows}']:
        for cell in row:
            cell.font = Font(bold=True,color='25A2FF')

    wb.save(f'{to_path}/{county}.xlsx')

